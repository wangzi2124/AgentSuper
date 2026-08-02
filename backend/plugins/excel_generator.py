"""
Excel Generator Plugin

Creates Excel spreadsheets (.xlsx) from structured content using openpyxl.
"""
import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PLUGIN_NAME = "excel-generator"
PLUGIN_VERSION = "0.1.0"
PLUGIN_DESCRIPTION = "Creates Excel spreadsheets (.xlsx) from structured content"


def tool_create_excel(sheets: str = "[]", output_path: str = "") -> str:
    """
    Create an Excel spreadsheet (.xlsx) with one or more sheets.

    Parameters:
    - sheets: JSON array of sheet objects. Each object has:
        {"name": "Sheet1", "headers": ["Col1", "Col2", ...], "rows": [["val1", "val2", ...], ...]}
        At least one sheet is required.
    - output_path: optional absolute path to save (auto-generated if empty)
    """
    base_dir = Path(__file__).resolve().parents[1]
    output_dir = base_dir / "data" / "generated"
    output_dir.mkdir(parents=True, exist_ok=True)

    if output_path:
        from app.plugins._output import resolve_output_path
        try:
            target = resolve_output_path(output_path, ".xlsx")
        except ValueError as e:
            return f"Error: {e}"
    else:
        from datetime import datetime
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        target = output_dir / f"spreadsheet_{ts}.xlsx"

    try:
        parsed_sheets = json.loads(sheets)
    except json.JSONDecodeError as e:
        return f"Error: invalid sheets JSON - {e}"

    if not parsed_sheets:
        return "Error: at least one sheet is required"

    wb = Workbook()

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Arial", size=10)
    cell_alignment = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    for idx, sheet_data in enumerate(parsed_sheets):
        name = sheet_data.get("name", f"Sheet{idx + 1}")
        headers = sheet_data.get("headers", [])
        rows = sheet_data.get("rows", [])

        if idx == 0:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(title=name)

        # Write headers
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(h))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                if row_idx % 2 == 1:
                    cell.fill = alt_fill

        # Auto-fit column widths
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1]))
            for row in rows:
                if col_idx <= len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    wb.save(str(target))
    return f"Excel created successfully: {target}"
